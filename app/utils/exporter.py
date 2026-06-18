"""
民航维修风险日报导出模块
支持 PDF 和 Excel 两种格式
"""
import os
from datetime import datetime, date
from typing import List, Dict, Any
from io import BytesIO

from PySide6.QtWidgets import QFileDialog, QMessageBox
from PySide6.QtCore import QObject, Signal

from app.db.database import Database
from app.models.schemas import RiskJob, ContractProject, FollowUp

# ---------- 可选依赖探测 ----------
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

try:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_JUSTIFY
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
    )
    REPORTLAB_OK = True
except ImportError:
    REPORTLAB_OK = False

# 常见中文字体路径（Windows）
_CN_FONT_CANDIDATES = [
    r"C:\Windows\Fonts\msyh.ttc",    # 微软雅黑
    r"C:\Windows\Fonts\simhei.ttf",   # 黑体
    r"C:\Windows\Fonts\simsun.ttc",   # 宋体
    r"C:\Windows\Fonts\simkai.ttf",   # 楷体
]


def _register_cn_font() -> str:
    """注册中文字体，返回字体名。失败返回 Helvetica"""
    if not REPORTLAB_OK:
        return "Helvetica"
    font_name = "CNFont"
    for p in _CN_FONT_CANDIDATES:
        if os.path.exists(p):
            try:
                pdfmetrics.registerFont(TTFont(font_name, p))
                return font_name
            except Exception:
                continue
    return "Helvetica"


RISK_LEVEL_COLORS = {
    "低风险": "#27ae60",
    "中风险": "#f39c12",
    "高风险": "#e67e22",
    "极高风险": "#c0392b",
}


class DailyReportExporter(QObject):
    """日报导出管理器（由 UI 调用）"""

    def __init__(self, db: Database, project: ContractProject, job_date: date,
                 jobs: List[RiskJob], parent=None):
        super().__init__(parent)
        self.db = db
        self.project = project
        self.job_date = job_date
        self.jobs = jobs
        self.parent = parent
        self._cn_font = _register_cn_font() if REPORTLAB_OK else ""

    # ========== UI 入口 ==========
    def export_pdf(self):
        if not REPORTLAB_OK:
            QMessageBox.warning(self.parent, "导出失败",
                                "缺少 reportlab 依赖，请先运行：pip install reportlab")
            return
        default_name = f"{self.project.name}_{self.job_date.isoformat()}_风险日报.pdf"
        path, _ = QFileDialog.getSaveFileName(
            self.parent, "导出 PDF 日报", default_name, "PDF 文件 (*.pdf)")
        if not path:
            return
        try:
            self._build_pdf(path)
            QMessageBox.information(self.parent, "导出成功", f"PDF 已保存至：\n{path}")
        except Exception as e:
            import traceback
            QMessageBox.critical(self.parent, "导出失败",
                                 f"生成 PDF 出错：\n{e}\n\n详细：\n{traceback.format_exc()}")

    def export_excel(self):
        if not OPENPYXL_OK:
            QMessageBox.warning(self.parent, "导出失败",
                                "缺少 openpyxl 依赖，请先运行：pip install openpyxl")
            return
        default_name = f"{self.project.name}_{self.job_date.isoformat()}_风险日报.xlsx"
        path, _ = QFileDialog.getSaveFileName(
            self.parent, "导出 Excel 日报", default_name, "Excel 文件 (*.xlsx)")
        if not path:
            return
        try:
            self._build_excel(path)
            QMessageBox.information(self.parent, "导出成功", f"Excel 已保存至：\n{path}")
        except Exception as e:
            import traceback
            QMessageBox.critical(self.parent, "导出失败",
                                 f"生成 Excel 出错：\n{e}\n\n详细：\n{traceback.format_exc()}")

    # ========== 公共数据准备 ==========
    def _summary(self) -> Dict[str, List[RiskJob]]:
        now = datetime.now()
        today_start = datetime.combine(self.job_date, datetime.min.time())
        today_end = datetime.combine(self.job_date, datetime.max.time())
        return {
            "client": [j for j in self.jobs
                       if j.need_client_safety_officer and j.status != "已关闭"],
            "close_today": [j for j in self.jobs
                            if j.status == "进行中"
                            and j.estimated_end_time
                            and today_start <= j.estimated_end_time <= today_end],
            "overdue": [j for j in self.jobs
                        if j.status == "进行中"
                        and j.estimated_end_time and j.estimated_end_time < now],
        }

    # ========== PDF 导出 ==========
    def _build_pdf(self, path: str):
        doc = SimpleDocTemplate(
            path, pagesize=landscape(A4),
            leftMargin=12 * mm, rightMargin=12 * mm,
            topMargin=12 * mm, bottomMargin=12 * mm,
            title=f"{self.project.name} 风险日报 - {self.job_date}",
            author="民航维修风险日报系统",
        )
        cn = self._cn_font
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "TitleCN", parent=styles["Title"], fontName=cn, fontSize=20,
            leading=26, textColor=colors.HexColor("#2c3e50"), alignment=TA_CENTER, spaceAfter=6,
        )
        sub_style = ParagraphStyle(
            "SubCN", parent=styles["Normal"], fontName=cn, fontSize=11,
            leading=16, textColor=colors.HexColor("#555555"), alignment=TA_CENTER, spaceAfter=10,
        )
        h2_style = ParagraphStyle(
            "H2CN", parent=styles["Heading2"], fontName=cn, fontSize=14,
            leading=20, textColor=colors.HexColor("#2980b9"), spaceBefore=10, spaceAfter=6,
        )
        body_style = ParagraphStyle(
            "BodyCN", parent=styles["Normal"], fontName=cn, fontSize=9.5,
            leading=14, alignment=TA_LEFT, wordWrap="CJK",
        )
        small_style = ParagraphStyle(
            "SmallCN", parent=body_style, fontSize=8.5, leading=12,
            textColor=colors.HexColor("#555555"),
        )

        story = []

        # 1. 标题栏
        story.append(Paragraph(f"民航维修现场风险日报", title_style))
        story.append(Paragraph(
            f"合同项目：{self.project.name}　·　合同编号：{self.project.contract_no or '-'}　·　"
            f"作业日期：{self.job_date.isoformat()}　·　生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M')}",
            sub_style,
        ))
        story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#2980b9")))

        # 2. 协调会摘要
        summary = self._summary()
        story.append(Paragraph("📌 协调会摘要", h2_style))
        coord_data = [
            [Paragraph("👤 需客户安全员到场", small_style),
             Paragraph("⏰ 预计今天关闭", small_style),
             Paragraph("🚨 超时未关闭", small_style)],
            [Paragraph("<br/>".join([
                f"• [{j.risk_level}] {j.work_type} - {j.work_location} ({j.aircraft_no})"
                for j in summary["client"]]) or "无", body_style),
             Paragraph("<br/>".join([
                f"• [{j.risk_level}] {j.work_type} - {j.work_location} ({j.aircraft_no})"
                f"　预计 {j.estimated_end_time.strftime('%H:%M') if j.estimated_end_time else '-'}"
                for j in summary["close_today"]]) or "无", body_style),
             Paragraph("<br/>".join([
                f"• [{j.risk_level}] {j.work_type} - {j.work_location} ({j.aircraft_no})"
                for j in summary["overdue"]]) or "无", body_style),
            ]
        ]
        t = Table(coord_data, colWidths=[doc.width / 3.0] * 3)
        t.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, -1), cn),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#8e44ad")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cccccc")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        story.append(t)
        story.append(Spacer(1, 4 * mm))

        # 3. 按状态分组的风险卡片
        for status in ["未开工", "进行中", "已关闭"]:
            jobs_of_status = [j for j in self.jobs if j.status == status]
            color = {"未开工": "#2980b9", "进行中": "#e67e22", "已关闭": "#27ae60"}[status]
            story.append(Paragraph(
                f"🔹 {status}（{len(jobs_of_status)}）", h2_style))
            story.append(HRFlowable(width="100%", thickness=0.6,
                                    color=colors.HexColor(color)))

            if not jobs_of_status:
                story.append(Paragraph("（无）", small_style))
                story.append(Spacer(1, 2 * mm))
                continue

            for job in jobs_of_status:
                try:
                    story.append(self._job_card_pdf(job, body_style, small_style, cn, color))
                except Exception as e:
                    err_txt = f"[导出本条作业失败] {job.work_type} - {job.work_location}：{e}"
                    story.append(Paragraph(err_txt, small_style))
                story.append(Spacer(1, 2 * mm))

        doc.build(story)

    def _job_card_pdf(self, job: RiskJob, body_style, small_style, cn: str,
                      status_color: str):
        level_bg = RISK_LEVEL_COLORS.get(job.risk_level, "#7f8c8d")
        header_bg = colors.HexColor("#f8f9fa")
        issue_txt = ("⚠️ " + job.issues) if job.issues else ""
        client_tag = "👤 需客户安全员到场" if job.need_client_safety_officer else ""
        overdue_txt = ""
        if job.status == "已关闭":
            if self.db.is_overdue_closed(job):
                overdue_txt = "🔴 超时关闭"
            else:
                overdue_txt = "🟢 按时关闭"
        elif job.status == "进行中" and job.estimated_end_time \
                and job.estimated_end_time < datetime.now():
            overdue_txt = "🚨 超时进行中"

        rows = []
        # 标题行
        rows.append([
            Paragraph(f"<b>{job.work_type}</b>　<span color='white'>{job.risk_level}</span>",
                      body_style),
            Paragraph(f"位置：{job.work_location or '-'}　　机型：{job.aircraft_no or '-'}",
                      small_style),
        ])
        rows.append([
            Paragraph(
                f"班组：{job.team or '-'}　　负责人：{job.team_leader or '-'}　　"
                f"许可证：{job.permit_status}　{' | ' + client_tag if client_tag else ''}"
                f"{' | ' + overdue_txt if overdue_txt else ''}", small_style),
            Paragraph(
                f"预计结束：{job.estimated_end_time.strftime('%Y-%m-%d %H:%M') if job.estimated_end_time else '-'}　　"
                f"实际结束：{job.actual_end_time.strftime('%Y-%m-%d %H:%M') if job.actual_end_time else '-'}",
                small_style),
        ])
        desc_lines = [
            f"<b>风险描述：</b>{job.description or '-'}",
            f"<b>隔离措施：</b>{job.isolation_measures or '-'}",
            (f"<b>问题提示：</b>{issue_txt}") if issue_txt else "",
            (f"<b>项目经理意见：</b>{job.pm_comments}") if job.pm_comments else "",
            (f"<b>关闭说明：</b>{job.close_remark}") if job.close_remark else "",
        ]
        rows.append([Paragraph("<br/>".join([x for x in desc_lines if x]), body_style), ""])

        tbl = Table(rows, colWidths=["*", "*"])
        style_cmds = [
            ("FONTNAME", (0, 0), (-1, -1), cn),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#dddddd")),
            ("BACKGROUND", (0, 0), (1, 0), header_bg),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("SPAN", (0, 2), (1, 2)),
            ("LINEBEFORE", (0, 0), (0, -1), 3, colors.HexColor(status_color)),
            # 风险等级色块（模拟：标题行背景+文字色）
        ]
        tbl.setStyle(TableStyle(style_cmds))

        # 跟进记录链
        story_parts = [tbl]
        follow_ups: List[FollowUp] = self.db.get_follow_ups_by_job(job.id or 0)
        if follow_ups:
            chain_data = [[
                Paragraph("时间", small_style),
                Paragraph("跟进人", small_style),
                Paragraph("跟进动作/复查安排", small_style),
                Paragraph("结果", small_style),
                Paragraph("闭环确认", small_style),
            ]]
            for fu in follow_ups:
                chain_data.append([
                    Paragraph(fu.follow_time.strftime('%m-%d %H:%M') if fu.follow_time else '-', small_style),
                    Paragraph(fu.owner or '-', small_style),
                    Paragraph(
                        (fu.action or '-') +
                        (f"<br/>复查日期：{fu.review_date.isoformat()}" if fu.review_date else ""),
                        small_style),
                    Paragraph(fu.result or '-', small_style),
                    Paragraph(("✅ " + fu.confirmed_by +
                               (f"<br/>{fu.confirmed_at.strftime('%m-%d')}" if fu.confirmed_at else ""))
                              if fu.confirmed else "⏳ 待确认", small_style),
                ])
            chain_tbl = Table(chain_data, colWidths=[18 * mm, 24 * mm, "*", "*", 24 * mm])
            chain_tbl.setStyle(TableStyle([
                ("FONTNAME", (0, 0), (-1, -1), cn),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#ecf0f1")),
                ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#cccccc")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]))
            story_parts.append(Spacer(1, 1 * mm))
            story_parts.append(Paragraph("📎 跟进链", small_style))
            story_parts.append(chain_tbl)

        container = Table([[p] for p in story_parts])
        container.setStyle(TableStyle([
            ("BOX", (0, 0), (-1, -1), 1, colors.HexColor("#bdc3c7")),
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ("TOPPADDING", (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ]))
        return container

    # ========== Excel 导出 ==========
    def _build_excel(self, path: str):
        wb = Workbook()

        # -------- Sheet1: 汇总 & 摘要 --------
        ws_summary = wb.active
        ws_summary.title = "日报汇总"

        bold = Font(bold=True)
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill("solid", fgColor="2980B9")
        coord_header_fill = PatternFill("solid", fgColor="8E44AD")
        thin = Side(border_style="thin", color="999999")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)

        # 标题
        ws_summary.merge_cells("A1:F1")
        c = ws_summary["A1"]
        c.value = f"民航维修现场风险日报 - {self.project.name}"
        c.font = Font(bold=True, size=16, color="2C3E50")
        c.alignment = center
        ws_summary.row_dimensions[1].height = 32

        # 基本信息
        info_lines = [
            ("合同编号", self.project.contract_no or "-", "作业日期",
             self.job_date.isoformat(), "生成时间",
             datetime.now().strftime("%Y-%m-%d %H:%M")),
            ("风险作业总数", str(len(self.jobs)),
             "未开工", str(sum(1 for j in self.jobs if j.status == "未开工")),
             "进行中", str(sum(1 for j in self.jobs if j.status == "进行中"))),
            ("已关闭", str(sum(1 for j in self.jobs if j.status == "已关闭")),
             "存在问题", str(sum(1 for j in self.jobs if j.issues)),
             "需客户安全员", str(sum(1 for j in self.jobs if j.need_client_safety_officer and j.status != "已关闭"))),
        ]
        start_row = 2
        for i, line in enumerate(info_lines):
            for j, val in enumerate(line):
                col = j + 1
                cell = ws_summary.cell(row=start_row + i, column=col, value=val)
                cell.border = border
                if j % 2 == 0:
                    cell.font = bold
                    cell.fill = PatternFill("solid", fgColor="F4F6F7")
                    cell.alignment = center
                else:
                    cell.alignment = center

        # 协调会摘要
        coord_row = start_row + len(info_lines) + 2
        ws_summary.merge_cells(start_row=coord_row, start_column=1,
                               end_row=coord_row, end_column=6)
        hcell = ws_summary.cell(row=coord_row, column=1, value="📌 协调会摘要")
        hcell.font = header_font
        hcell.fill = coord_header_fill
        hcell.alignment = center

        summary = self._summary()
        coord_headers = ["👤 需客户安全员到场", "⏰ 预计今天关闭", "🚨 超时未关闭"]
        coord_lists = [summary["client"], summary["close_today"], summary["overdue"]]

        hrow = coord_row + 1
        for i, h in enumerate(coord_headers):
            cc = ws_summary.cell(row=hrow, column=1 + i * 2, value=h)
            cc.font = Font(bold=True, color="FFFFFF")
            cc.fill = PatternFill("solid", fgColor="9B59B6")
            cc.alignment = center
            cc.border = border
            ws_summary.merge_cells(start_row=hrow, start_column=1 + i * 2,
                                   end_row=hrow, end_column=2 + i * 2)

        max_len = max(len(lst) for lst in coord_lists) if coord_lists else 1
        for r in range(max_len):
            row_idx = hrow + 1 + r
            for i, lst in enumerate(coord_lists):
                col_start = 1 + i * 2
                if r < len(lst):
                    j = lst[r]
                    text = (f"[{j.risk_level}] {j.work_type}\n"
                            f"位置：{j.work_location} 机型：{j.aircraft_no}\n"
                            f"班组：{j.team} 负责人：{j.team_leader}"
                            f"{'\n预计：' + j.estimated_end_time.strftime('%H:%M') if j.estimated_end_time else ''}")
                else:
                    text = "—"
                cell = ws_summary.cell(row=row_idx, column=col_start, value=text)
                cell.alignment = left_wrap
                cell.border = border
                ws_summary.merge_cells(start_row=row_idx, start_column=col_start,
                                       end_row=row_idx, end_column=col_start + 1)

        # 列宽
        for col_idx, w in enumerate([14, 14, 14, 14, 14, 14], 1):
            ws_summary.column_dimensions[get_column_letter(col_idx)].width = w

        # -------- Sheet2: 风险作业明细 --------
        ws = wb.create_sheet("风险作业明细")
        detail_headers = [
            "编号", "状态", "风险等级", "作业类型", "作业位置", "机型",
            "班组", "负责人", "许可证状态", "许可证到期",
            "参与人员", "预计结束", "实际结束", "关闭结果",
            "需客户安全员", "风险描述", "隔离措施",
            "问题提示", "项目经理意见", "关闭说明",
        ]
        for col_idx, h in enumerate(detail_headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border

        personnel_name_cache = {}

        def pname(pid):
            if pid not in personnel_name_cache:
                p = self.db.get_personnel_by_ids([pid])
                personnel_name_cache[pid] = p[0].name if p else f"#{pid}"
            return personnel_name_cache[pid]

        for row_idx, job in enumerate(self.jobs, 2):
            close_result = "-"
            if job.status == "已关闭":
                close_result = "超时关闭" if self.db.is_overdue_closed(job) else "按时关闭"
            values = [
                job.id,
                job.status,
                job.risk_level,
                job.work_type,
                job.work_location,
                job.aircraft_no,
                job.team,
                job.team_leader,
                job.permit_status,
                job.permit_expiry.isoformat() if job.permit_expiry else "",
                "、".join([pname(pid) for pid in job.personnel_ids]),
                job.estimated_end_time.strftime("%Y-%m-%d %H:%M") if job.estimated_end_time else "",
                job.actual_end_time.strftime("%Y-%m-%d %H:%M") if job.actual_end_time else "",
                close_result,
                "是" if job.need_client_safety_officer else "否",
                job.description,
                job.isolation_measures,
                job.issues,
                job.pm_comments,
                job.close_remark,
            ]
            for col_idx, v in enumerate(values, 1):
                cc = ws.cell(row=row_idx, column=col_idx, value=v)
                cc.alignment = left_wrap
                cc.border = border
                # 风险等级着色
                if col_idx == 3 and v in RISK_LEVEL_COLORS:
                    cc.font = Font(bold=True, color="FFFFFF")
                    cc.fill = PatternFill("solid", fgColor=RISK_LEVEL_COLORS[v].lstrip("#"))
                    cc.alignment = center
                if col_idx == 2:
                    status_color = {"未开工": "2980B9", "进行中": "E67E22",
                                    "已关闭": "27AE60"}.get(v, "7F8C8D")
                    cc.font = Font(bold=True, color="FFFFFF")
                    cc.fill = PatternFill("solid", fgColor=status_color)
                    cc.alignment = center
                if col_idx == 14:
                    if v == "按时关闭":
                        cc.fill = PatternFill("solid", fgColor="27AE60")
                        cc.font = Font(bold=True, color="FFFFFF")
                        cc.alignment = center
                    elif v == "超时关闭":
                        cc.fill = PatternFill("solid", fgColor="C0392B")
                        cc.font = Font(bold=True, color="FFFFFF")
                        cc.alignment = center

        # 列宽
        widths = [6, 9, 10, 12, 18, 12, 12, 10, 11, 12, 24, 18, 18, 10, 13, 40, 40, 30, 30, 40]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A2"

        # -------- Sheet3: 跟进记录明细 --------
        ws_fu = wb.create_sheet("跟进记录")
        fu_headers = [
            "作业编号", "作业类型", "作业位置", "机型", "状态",
            "跟进时间", "跟进人", "跟进动作", "复查日期",
            "复查结果", "闭环确认", "确认人", "确认时间"
        ]
        for col_idx, h in enumerate(fu_headers, 1):
            c = ws_fu.cell(row=1, column=col_idx, value=h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = center
            c.border = border

        job_cache = {j.id: j for j in self.jobs}
        all_fus: List[FollowUp] = []
        for j in self.jobs:
            all_fus.extend(self.db.get_follow_ups_by_job(j.id or 0))

        for row_idx, fu in enumerate(all_fus, 2):
            job = job_cache.get(fu.job_id)
            vals = [
                fu.job_id,
                job.work_type if job else "-",
                job.work_location if job else "-",
                job.aircraft_no if job else "-",
                job.status if job else "-",
                fu.follow_time.strftime("%Y-%m-%d %H:%M") if fu.follow_time else "-",
                fu.owner or "-",
                fu.action or "-",
                fu.review_date.isoformat() if fu.review_date else "-",
                fu.result or "-",
                "✅ 已确认" if fu.confirmed else "⏳ 待确认",
                fu.confirmed_by or "-",
                fu.confirmed_at.strftime("%Y-%m-%d %H:%M") if fu.confirmed_at else "-",
            ]
            for col_idx, v in enumerate(vals, 1):
                cc = ws_fu.cell(row=row_idx, column=col_idx, value=v)
                cc.alignment = left_wrap
                cc.border = border

        widths = [8, 12, 18, 12, 9, 17, 12, 36, 12, 30, 12, 12, 17]
        for i, w in enumerate(widths, 1):
            ws_fu.column_dimensions[get_column_letter(i)].width = w
        ws_fu.freeze_panes = "A2"

        wb.save(path)
